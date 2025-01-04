#!/usr/bin/env python3

import os
import sys
import time
import json
import re
import logging

import requests
import inquirer
from tqdm import tqdm
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

# -------------------------------------------------------------------
# Logging configuration
# -------------------------------------------------------------------
logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(message)s')

stdout_handler = logging.StreamHandler(sys.stdout)
stdout_handler.setLevel(logging.INFO)
stdout_handler.setFormatter(formatter)

logger.addHandler(stdout_handler)

# -------------------------------------------------------------------
# Cloud parameters
# -------------------------------------------------------------------
cloud_email = "rodolfobortolin@gmail.com"
cloud_token = ""
cloud_base_URL = "https://<domain>.atlassian.net"

# -------------------------------------------------------------------
# Global workbook and sheets (will be loaded after the initial checks)
# -------------------------------------------------------------------
wBook = None
wScheetUsers = None
wScheetCustomFields = None
wScheetProjects = None
wScheetStatus = None
wScheetPriority = None
wScheetIssueType = None

# -------------------------------------------------------------------
# This list will store all mappings: serverID => cloudID.
# We will generate a final Excel (mapping_result.xlsx) from it.
# -------------------------------------------------------------------
mapping_data = []

# -------------------------------------------------------------------
# Function to test Jira Cloud connection
# -------------------------------------------------------------------
def test_jira_cloud_connection():
    """
    Tests connection to Jira Cloud by calling the 'myself' endpoint.
    Returns True if status_code == 200, otherwise False.
    Reference: https://developer.atlassian.com/cloud/jira/platform/rest/v3/#api-rest-api-3-myself-get
    """
    try:
        test_url = f"{cloud_base_URL}/rest/api/3/myself"
        response = requests.get(test_url, auth=HTTPBasicAuth(cloud_email, cloud_token))
        if response.status_code == 200:
            return True
        else:
            logger.info(f"[Connection Test] Unexpected status code: {response.status_code}")
            return False
    except Exception as e:
        logger.info(f"[Connection Test] Exception: {str(e)}")
        return False

# -------------------------------------------------------------------
# Initial requirement checks
# -------------------------------------------------------------------
def initial_checks():
    """
    1. Check if 'mapping.xlsx' exists.
    2. Check Jira Cloud connection.
    3. Check if there is at least one JSON file containing a 'rules' key.
    """
    # 1) Check if mapping.xlsx exists
    if not os.path.isfile('mapping.xlsx'):
        logger.info("\n❌ 'mapping.xlsx' not found in the current directory.")
        return False
    
    # 2) Check Jira Cloud connection
    logger.info("\nTesting Jira Cloud connection...")
    if not test_jira_cloud_connection():
        logger.info("❌ Unable to connect to Jira Cloud. Please verify your credentials and network.")
        return False
    logger.info("✔ Jira Cloud connection OK.")
    
    # 3) Check if there is at least one JSON file containing automation rules
    logger.info("\nChecking for at least one JSON file with 'rules' in it...")
    json_files = [f for f in os.listdir('.') if f.endswith('.json')]
    if not json_files:
        logger.info("❌ No JSON files found in the current directory.")
        return False

    found_rules = False
    for jf in json_files:
        try:
            with open(jf, 'r', encoding='utf-8') as file_data:
                content = json.load(file_data)
                if "rules" in content:
                    found_rules = True
                    break
        except Exception:
            continue
    
    if not found_rules:
        logger.info("❌ Could not find any JSON file containing 'rules'.")
        return False
    
    logger.info("✔ Found at least one JSON file containing 'rules'.")
    return True

# -------------------------------------------------------------------
# Helper methods (cloud queries, Excel lookups, etc.)
# -------------------------------------------------------------------
def getEmailforUserInExcel(username):
    row = 0
    for cell in wScheetUsers['A']:
        row += 1
        if cell.value is not None and username == str(cell.value).strip():
            return wScheetUsers["B" + str(row)].value
    return None

def getCustomFieldNameInExcel(cf_id):
    row = 0
    for cell in wScheetCustomFields['A']:
        row += 1
        if cell.value is not None:
            # Must match exactly "customfield_XX"
            if cf_id == "customfield_" + str(cell.value).strip():
                return wScheetCustomFields["B" + str(row)].value
    return None

def getStatusNameInExcel(id):
    row = 0
    for cell in wScheetStatus['A']:
        row += 1
        if cell.value is not None and str(cell.value).strip() == str(id).strip():
            return wScheetStatus["B" + str(row)].value
    return None

def getIssueTypeNameInExcel(id):
    row = 0
    for cell in wScheetIssueType['A']:
        row += 1
        if cell.value is not None and str(cell.value).strip() == str(id).strip():
            return wScheetIssueType["B" + str(row)].value
    return None

def getPriorityNameInExcel(id):
    row = 0
    for cell in wScheetPriority['A']:
        row += 1
        if cell.value is not None and str(cell.value).strip() == str(id).strip():
            return wScheetPriority["B" + str(row)].value
    return None

# Cloud lookups
def getCustomFieldIdInCloud(name):
    response = requests.get(
        f"{cloud_base_URL}/rest/api/3/field/search?query={name}",
        auth=HTTPBasicAuth(cloud_email, cloud_token),
    )
    if response.status_code != 200:
        return None
    customfields = response.json()
    for cf in customfields['values']:
        if cf['name'] == name:
            return cf['id']
    return None

def getStatusIdInCloud(name):
    if not name:
        return None
    response = requests.get(
        f"{cloud_base_URL}/rest/api/3/statuses/search?searchString={name}",
        auth=HTTPBasicAuth(cloud_email, cloud_token),
    )
    if response.status_code != 200:
        return None
    statuses = response.json()
    for status in statuses['values']:
        if status['name'] == name:
            return status['id']
    return None

def getPriorityIdInCloud(name):
    if not name:
        return None
    response = requests.get(
        f"{cloud_base_URL}/rest/api/3/priority/search",
        auth=HTTPBasicAuth(cloud_email, cloud_token),
    )
    if response.status_code != 200:
        return None
    priorities = response.json()
    for priority in priorities['values']:
        if priority['name'] == name:
            return priority['id']
    return None

def getIssueTypeIdInCloud(name):
    if not name:
        return None
    response = requests.get(
        f"{cloud_base_URL}/rest/api/3/issuetype",
        auth=HTTPBasicAuth(cloud_email, cloud_token),
    )
    if response.status_code != 200:
        return None
    issueTypes = response.json()
    for issueType in issueTypes:
        if issueType['name'] == name:
            return issueType['id']
    return None

def getAccountIdInCloud(email):
    if not email:
        return None
    response = requests.get(
        f"{cloud_base_URL}/rest/api/3/user/search?query={email}",
        auth=HTTPBasicAuth(cloud_email, cloud_token),
    )
    if response.status_code != 200:
        return None
    json_user = response.json()
    if len(json_user) > 0:
        return json_user[0]['accountId']
    return None

def getProjectIdInCloud(projectKey):
    if not projectKey:
        return None
    response = requests.get(
        f"{cloud_base_URL}/rest/api/3/project/search?keys={projectKey}",
        auth=HTTPBasicAuth(cloud_email, cloud_token),
    )
    if response.status_code != 200:
        return None
    projects = response.json()
    for project in projects['values']:
        if project['key'] == projectKey:
            return project['id']
    return None

# -------------------------------------------------------------------
# Class with all cloud-related transformation functions
# -------------------------------------------------------------------
class cloud:

    @staticmethod
    def removeDisabled(filename):
        """
        Removes disabled rules from the JSON automation file to keep only ENABLED ones.
        """
        with open(filename, 'r', encoding='utf-8') as f:
            my_list = json.load(f)

        with open(filename + "-original-pretty.json", 'w', encoding='utf-8') as out_json_file:
            json.dump(my_list, out_json_file, indent=2)

        automations = my_list.get('rules', [])
        enableList = [rule for rule in automations if rule.get('state') == 'ENABLED']

        obj = {"rules": enableList, "cloud": False}

        with open(filename + "-modified-for-cloud.json", 'w', encoding='utf-8') as out_json_file:
            json.dump(obj, out_json_file)

    @staticmethod
    def replaceFixedFields(fileName):
        """
        Replaces certain known references or strings (like 'Customer Request Type') with their Cloud counterparts.
        Also changes references to customfield_<ID> from type=ID to type=NAME for the next processing steps.
        """
        with open(fileName + "-modified-for-cloud.json", "rt", encoding='utf8') as fin:
            data = fin.read()

        cfsInSheet = []
        for row, cell in enumerate(wScheetCustomFields['A'], start=1):
            if cell.value and str(cell.value).strip():
                cfsInSheet.append(str(cell.value).strip())

        replaced_count = 0
        for cf_id in tqdm(cfsInSheet, desc="Replacing fixed custom fields", ncols=100):
            old_str = f'type": "ID", "value": "customfield_{cf_id}'
            new_name = getCustomFieldNameInExcel("customfield_" + cf_id)
            if new_name and old_str in data:
                new_str = f'type": "NAME", "value": "{new_name}'
                data = data.replace(old_str, new_str)
                replaced_count += 1

        # "Customer Request Type" -> "Request Type"
        if "Customer Request Type" in data:
            data = data.replace('Customer Request Type', 'Request Type')
            replaced_count += 1

        with open(fileName + "-modified-for-cloud.json", "wt", encoding='utf8') as fout:
            fout.write(data)

        logger.info(f"\n[Summary] Fixed Fields: {replaced_count} replacements done.")

    @staticmethod
    def replaceCustomFields(fileName):
        """
        Searches for customfield_<ID> references and replaces them with their corresponding new Cloud field IDs.
        """
        with open(fileName + "-modified-for-cloud.json", "rt", encoding='utf8') as fin:
            data = fin.read()

        occurrences = list(re.finditer("customfield_", data))
        missingCustomFields = []
        replaced_count = 0

        with tqdm(total=len(occurrences), desc="Replacing custom fields", ncols=100) as pbar:
            for match in occurrences:
                pbar.update(1)
                oldId = data[match.start():match.end() + 5]  # e.g. 'customfield_10000'
                cfName = getCustomFieldNameInExcel(oldId)

                # If we can't match a name from Excel, skip adding to mapping_data
                if not cfName:
                    continue

                newId = getCustomFieldIdInCloud(cfName)
                if newId:
                    data = data.replace(oldId, newId)
                    replaced_count += 1
                    mapping_data.append({
                        "type": "customfield",
                        "name": cfName,
                        "server_id": oldId,
                        "cloud_id": newId
                    })
                else:
                    missingCustomFields.append(cfName)
                    mapping_data.append({
                        "type": "customfield",
                        "name": cfName,
                        "server_id": oldId,
                        "cloud_id": None
                    })

        with open(fileName + "-modified-for-cloud.json", "wt", encoding='utf8') as fout:
            fout.write(data)

        logger.info(f"\n[Summary] Custom Fields replaced: {replaced_count}. Missing: {len(missingCustomFields)}")
        if missingCustomFields:
            logger.info(f"Missing in Cloud: {missingCustomFields}")

    @staticmethod
    def replaceStatus(fileName, templates):
        """
        Replaces status IDs based on Excel sheet. 
        E.g., "status":"10100" => "status":"<cloud id>"
        """
        with open(fileName + "-modified-for-cloud.json", "rt", encoding='utf8') as fin:
            data = fin.read()

        missingStatuses = []
        replaced_count = 0

        status_rows = []
        row_index = 1
        for cell in wScheetStatus['A']:
            row_index += 1
            if cell.value and str(cell.value).strip():
                status_rows.append(row_index)

        with tqdm(total=len(status_rows), desc="Replacing status IDs", ncols=100) as pbar:
            for row in status_rows:
                pbar.update(1)
                original_id = str(wScheetStatus["A" + str(row)].value).strip()
                name = wScheetStatus["B" + str(row)].value
                if not name:
                    continue

                cloudStatusId = getStatusIdInCloud(name)
                if cloudStatusId:
                    for template in templates:
                        fromStr = template + original_id + "\""
                        toStr = template + str(cloudStatusId) + "\""
                        if fromStr in data:
                            data = data.replace(fromStr, toStr)
                            replaced_count += 1
                    mapping_data.append({
                        "type": "status",
                        "name": name,
                        "server_id": original_id,
                        "cloud_id": cloudStatusId
                    })
                else:
                    missingStatuses.append(name)
                    mapping_data.append({
                        "type": "status",
                        "name": name,
                        "server_id": original_id,
                        "cloud_id": None
                    })

        with open(fileName + "-modified-for-cloud.json", "wt", encoding='utf8') as fout:
            fout.write(data)

        logger.info(f"\n[Summary] Status replaced: {replaced_count}, missing in Cloud: {len(missingStatuses)}")
        if missingStatuses:
            logger.info(f"Missing statuses: {missingStatuses}")

    @staticmethod
    def replaceUsers(fileName, templates):
        """
        Replaces user references (server user keys) with the corresponding Cloud account IDs.
        Note: 
        You can use a separate script (e.g., generate-mappings.py) to generate the mapping data for up to 1000 users.
        For more extensive user sets, retrieving data directly from the DB is recommended.
        """
        with open(fileName + "-modified-for-cloud.json", "rt", encoding='utf8') as fin:
            data = fin.read()

        usersInSheet = []
        for cell in wScheetUsers['A']:
            val = cell.value
            if val and str(val).strip():
                usersInSheet.append(str(val).strip())

        usersInJson = []
        for userInSheet in usersInSheet:
            if userInSheet in data:
                usersInJson.append(userInSheet)

        replaced_count = 0
        missing_count = 0
        with tqdm(total=len(usersInJson), desc="Replacing users", ncols=100) as pbar:
            for user in usersInJson:
                pbar.update(1)
                userEmail = getEmailforUserInExcel(user)
                if not userEmail:
                    continue
                accountId = getAccountIdInCloud(userEmail)
                if accountId:
                    any_replaced = False
                    for template in templates:
                        fromStr = template + user + "\""
                        toStr = template + accountId + "\""
                        if fromStr in data:
                            data = data.replace(fromStr, toStr)
                            replaced_count += 1
                            any_replaced = True
                    if any_replaced:
                        mapping_data.append({
                            "type": "user",
                            "name": userEmail,
                            "server_id": user,
                            "cloud_id": accountId
                        })
                else:
                    missing_count += 1
                    mapping_data.append({
                        "type": "user",
                        "name": userEmail,
                        "server_id": user,
                        "cloud_id": None
                    })

        with open(fileName + "-modified-for-cloud.json", "wt", encoding='utf8') as fout:
            fout.write(data)

        logger.info(f"\n[Summary] Users replaced: {replaced_count}, missing in Cloud: {missing_count}")

    @staticmethod
    def replaceJIRAUSERUsers(fileName):
        """
        Specifically replaces 'JIRAUSERxxxx' references in the rules with the correct Cloud account IDs.
        Note: 
        You can use a separate script (e.g., generate-mappings.py) to generate user data for up to 1000 users.
        For more extensive user sets, retrieving data directly from the DB is recommended.
        """
        with open(fileName + "-modified-for-cloud.json", "rt", encoding='utf8') as fin:
            data = fin.read()

        jiraUsersInSheet = []
        for cell in wScheetUsers['A']:
            val = cell.value
            if val and str(val).strip().startswith("JIRAUSER"):
                jiraUsersInSheet.append(str(val).strip())

        usersInJson = []
        for user in jiraUsersInSheet:
            if user in data:
                usersInJson.append(user)

        replaced_count = 0
        missing_count = 0
        with tqdm(total=len(usersInJson), desc="Replacing JIRAUSER references", ncols=100) as pbar:
            for user in usersInJson:
                pbar.update(1)
                userEmail = getEmailforUserInExcel(user)
                if not userEmail:
                    continue
                accountId = getAccountIdInCloud(userEmail)
                if accountId:
                    fromStr = user + "\""
                    toStr = accountId + "\""
                    if fromStr in data:
                        data = data.replace(fromStr, toStr)
                        replaced_count += 1
                    mapping_data.append({
                        "type": "user",
                        "name": userEmail,
                        "server_id": user,
                        "cloud_id": accountId
                    })
                else:
                    missing_count += 1
                    mapping_data.append({
                        "type": "user",
                        "name": userEmail,
                        "server_id": user,
                        "cloud_id": None
                    })

        with open(fileName + "-modified-for-cloud.json", "wt", encoding='utf8') as fout:
            fout.write(data)

        logger.info(f"\n[Summary] JIRAUSER replaced: {replaced_count}, missing in Cloud: {missing_count}")

    @staticmethod
    def replacePriority(fileName, templates):
        """
        Replaces priority IDs based on Excel sheet.
        E.g., "priority":"2" => "priority":"<cloud id>"
        """
        with open(fileName + "-modified-for-cloud.json", "rt", encoding='utf8') as fin:
            data = fin.read()

        missingPriorities = []
        replaced_count = 0

        priority_rows = []
        row_index = 1
        for cell in wScheetPriority['A']:
            row_index += 1
            if cell.value and str(cell.value).strip():
                priority_rows.append(row_index)

        with tqdm(total=len(priority_rows), desc="Replacing priorities", ncols=100) as pbar:
            for row in priority_rows:
                pbar.update(1)
                original_id = str(wScheetPriority["A" + str(row)].value).strip()
                name = wScheetPriority["B" + str(row)].value
                if not name:
                    continue

                cloudPriorityId = getPriorityIdInCloud(name)
                if cloudPriorityId:
                    for template in templates:
                        fromStr = template + original_id + "\""
                        toStr = template + str(cloudPriorityId) + "\""
                        if fromStr in data:
                            data = data.replace(fromStr, toStr)
                            replaced_count += 1
                    mapping_data.append({
                        "type": "priority",
                        "name": name,
                        "server_id": original_id,
                        "cloud_id": cloudPriorityId
                    })
                else:
                    missingPriorities.append(name)
                    mapping_data.append({
                        "type": "priority",
                        "name": name,
                        "server_id": original_id,
                        "cloud_id": None
                    })

        with open(fileName + "-modified-for-cloud.json", "wt", encoding='utf8') as fout:
            fout.write(data)

        logger.info(
            f"\n[Summary] Priorities replaced: {replaced_count}, missing in Cloud: {len(missingPriorities)}"
        )
        if missingPriorities:
            logger.info(f"Missing priorities: {missingPriorities}")

    @staticmethod
    def replaceIssueType(fileName, templates):
        """
        Replaces issue type IDs based on Excel sheet.
        E.g., "issuetype":"1" => "issuetype":"<cloud id>"
        """
        with open(fileName + "-modified-for-cloud.json", "rt", encoding='utf8') as fin:
            data = fin.read()

        missingIssueTypes = []
        replaced_count = 0

        issuetype_rows = []
        row_index = 1
        for cell in wScheetIssueType['A']:
            row_index += 1
            if cell.value and str(cell.value).strip():
                issuetype_rows.append(row_index)

        with tqdm(total=len(issuetype_rows), desc="Replacing issue types", ncols=100) as pbar:
            for row in issuetype_rows:
                pbar.update(1)
                original_id = str(wScheetIssueType["A" + str(row)].value).strip()
                name = wScheetIssueType["B" + str(row)].value
                if not name:
                    continue

                cloudIssueTypeId = getIssueTypeIdInCloud(name)
                if cloudIssueTypeId:
                    for template in templates:
                        fromStr = template + original_id + "\""
                        toStr = template + str(cloudIssueTypeId) + "\""
                        if fromStr in data:
                            data = data.replace(fromStr, toStr)
                            replaced_count += 1
                    mapping_data.append({
                        "type": "issuetype",
                        "name": name,
                        "server_id": original_id,
                        "cloud_id": cloudIssueTypeId
                    })
                else:
                    missingIssueTypes.append(name)
                    mapping_data.append({
                        "type": "issuetype",
                        "name": name,
                        "server_id": original_id,
                        "cloud_id": None
                    })

        with open(fileName + "-modified-for-cloud.json", "wt", encoding='utf8') as fout:
            fout.write(data)

        logger.info(
            f"\n[Summary] Issue Types replaced: {replaced_count}, missing in Cloud: {len(missingIssueTypes)}"
        )
        if missingIssueTypes:
            logger.info(f"Missing issue types: {missingIssueTypes}")

    @staticmethod
    def replaceProject(fileName, templates):
        """
        Replaces project references (IDs) with their Cloud counterparts,
        based on the 'projects' Excel sheet.
        """
        with open(fileName + "-modified-for-cloud.json", "rt", encoding='utf8') as fin:
            data = fin.read()

        missingProjects = []
        replaced_count = 0

        project_rows = []
        row_index = 1
        for cell in wScheetProjects['A']:
            row_index += 1
            if cell.value and str(cell.value).strip():
                project_rows.append(row_index)

        with tqdm(total=len(project_rows), desc="Replacing projects", ncols=100) as pbar:
            for row in project_rows:
                pbar.update(1)
                original_id = str(wScheetProjects["A" + str(row)].value).strip()
                key = wScheetProjects["B" + str(row)].value
                if not key:
                    continue

                cloudProjectId = getProjectIdInCloud(key)
                if cloudProjectId:
                    for template in templates:
                        fromStr = template + original_id + "\""
                        toStr = template + str(cloudProjectId) + "\""
                        if fromStr in data:
                            data = data.replace(fromStr, toStr)
                            replaced_count += 1
                    mapping_data.append({
                        "type": "project",
                        "name": key,
                        "server_id": original_id,
                        "cloud_id": cloudProjectId
                    })
                else:
                    missingProjects.append(key)
                    mapping_data.append({
                        "type": "project",
                        "name": key,
                        "server_id": original_id,
                        "cloud_id": None
                    })

        with open(fileName + "-modified-for-cloud.json", "wt", encoding='utf8') as fout:
            fout.write(data)

        logger.info(
            f"\n[Summary] Projects replaced: {replaced_count}, missing in Cloud: {len(missingProjects)}"
        )
        if missingProjects:
            logger.info(f"Missing projects: {missingProjects}")

    @staticmethod
    def replaceOneOfCondition(fileName, fieldType):
        """
        Handles the 'ONE_OF' or 'NOT_ONE_OF' conditions for a given fieldType (status, issuetype, or priority).
        Example usage: cloud.replaceOneOfCondition(filename, 'status').
        """
        with open(fileName + "-modified-for-cloud.json", "rt", encoding='utf8') as fin:
            data = fin.read()

        patterns = [
            f'"selectedFieldType": "{fieldType}", "comparison": "ONE_OF", "compareValue": {{"type": "ID", "value": "(\\[.*?\\])"',
            f'"selectedFieldType": "{fieldType}", "comparison": "NOT_ONE_OF", "compareValue": {{"type": "ID", "value": "(\\[.*?\\])"'
        ]

        replaced_count = 0
        with tqdm(total=len(patterns), desc=f"Replacing ONE_OF/NOT_ONE_OF for {fieldType}", ncols=100) as pbar:
            for pattern in patterns:
                pbar.update(1)
                matches = list(re.finditer(pattern, data))
                for match in matches:
                    id_array = match.group(1)
                    clean_array = (
                        id_array.replace('\\"', '"')
                               .replace('"[', '[')
                               .replace(']"', ']')
                    )
                    id_list = json.loads(clean_array)

                    new_ids = []
                    for old_id in id_list:
                        old_id = old_id.strip()
                        if not old_id:
                            continue
                        new_id = None
                        if fieldType == 'status':
                            name = getStatusNameInExcel(old_id)
                            new_id = getStatusIdInCloud(name) if name else None
                            if name or new_id:
                                mapping_data.append({
                                    "type": "status",
                                    "name": name,
                                    "server_id": old_id,
                                    "cloud_id": new_id
                                })
                        elif fieldType == 'issuetype':
                            name = getIssueTypeNameInExcel(old_id)
                            new_id = getIssueTypeIdInCloud(name) if name else None
                            if name or new_id:
                                mapping_data.append({
                                    "type": "issuetype",
                                    "name": name,
                                    "server_id": old_id,
                                    "cloud_id": new_id
                                })
                        elif fieldType == 'priority':
                            name = getPriorityNameInExcel(old_id)
                            new_id = getPriorityIdInCloud(name) if name else None
                            if name or new_id:
                                mapping_data.append({
                                    "type": "priority",
                                    "name": name,
                                    "server_id": old_id,
                                    "cloud_id": new_id
                                })

                        if new_id:
                            new_ids.append(str(new_id))

                    # Replace the old array with the new array of IDs
                    old_value = f'"value": "[\\"{"\\\",\\\"".join(id_list)}\\"]"'
                    new_value = f'"value": "[\\"{"\\\",\\\"".join(new_ids)}\\"]"'
                    if old_value in data:
                        data = data.replace(old_value, new_value)
                        replaced_count += 1

        with open(fileName + "-modified-for-cloud.json", "wt", encoding='utf8') as fout:
            fout.write(data)

        logger.info(f"\n[Summary] {fieldType} ONE_OF/NOT_ONE_OF replacements: {replaced_count}")

    @staticmethod
    def formatJSON(filename):
        """
        Creates a pretty-printed JSON file for final consumption.
        """
        with open(filename + "-modified-for-cloud.json", 'r', encoding='utf-8') as fin:
            content = json.load(fin)

        with open(filename + "-modified-for-cloud-pretty.json", 'w', encoding='utf-8') as out_json_file:
            json.dump(content, out_json_file, indent=2)

    @staticmethod
    def generateMappingExcel(mapping_data, output_file="mapping_result.xlsx"):
        """
        Create a new Excel file with columns:
        A: Type
        B: Name
        C: Server ID
        D: Cloud ID
        E: Missing? (highlighted in red if missing)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Mappings"

        # Header
        headers = ["Type", "Name", "Server ID", "Cloud ID", "Missing?"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

        row = 2
        for entry in mapping_data:
            # If both server_id and name are None, skip writing
            if (not entry["server_id"] or entry["server_id"] == "None") and \
               (not entry["name"] or entry["name"] == "None"):
                continue

            ws.cell(row=row, column=1, value=entry["type"])
            ws.cell(row=row, column=2, value=entry["name"])
            ws.cell(row=row, column=3, value=str(entry["server_id"]))
            ws.cell(row=row, column=4, value=str(entry["cloud_id"]) if entry["cloud_id"] else None)

            missing_cell = ws.cell(row=row, column=5)
            if entry["cloud_id"] is None:
                missing_cell.value = "YES"
                missing_cell.fill = redFill
            else:
                missing_cell.value = "NO"

            row += 1

        wb.save(output_file)
        logger.info(f"\n[Summary] Generated mapping file: {output_file}")

# -------------------------------------------------------------------
# Functions for user interaction
# -------------------------------------------------------------------
def list_json_files():
    return [f for f in os.listdir('.') if f.endswith('.json')]

def select_file():
    json_files = list_json_files()
    if not json_files:
        print("\n❌ No JSON files found in the current directory!")
        return None

    questions = [
        inquirer.List(
            'filename',
            message='Select the Jira automation file',
            choices=json_files,
            carousel=True
        ),
        inquirer.Confirm(
            'split_files',
            message='Do you want to create separate files for each rule?',
            default=False
        )
    ]

    answers = inquirer.prompt(questions)
    return answers

def main():
    print("\nStarting Jira Server-to-Cloud automation migration...")

    # Perform the initial requirement checks
    if not initial_checks():
        print("\n❌ Missing requirements. Please fix the issues above and retry.")
        sys.exit(1)

    # Since the workbook exists, we can safely load it now
    global wBook, wScheetUsers, wScheetCustomFields, wScheetProjects
    global wScheetStatus, wScheetPriority, wScheetIssueType

    wBook = load_workbook('mapping.xlsx')
    wScheetUsers = wBook['users']
    wScheetCustomFields = wBook['customFields']
    wScheetProjects = wBook['projects']
    wScheetStatus = wBook['status']
    wScheetPriority = wBook['priority']
    wScheetIssueType = wBook['issuetype']

    # Note regarding generate-mappings.py
    print("\nNOTE: You can use the generate-mappings.py script to produce partial mappings for up to 1000 users. "
          "For larger user sets, it is recommended that you retrieve the data directly from your database.")

    # Request file selection
    answers = select_file()
    if not answers:
        return

    filename = answers['filename']
    splitFiles = answers['split_files']

    print(f"\nSelected file: {filename}")
    print("Settings:")
    print(f"   - Create separate files: {'Yes' if splitFiles else 'No'}")

    # Confirmation to proceed
    if not inquirer.confirm("Do you want to proceed with the migration?", default=True):
        print("\nOperation cancelled by user.")
        return

    print("\nStarting migration process...")

    try:
        # 1. Disable inactive rules
        print("\n1) Disabling inactive rules...")
        cloud.removeDisabled(filename)
        time.sleep(1)

        # 2. Replace fixed fields
        print("2) Replacing fixed custom fields...")
        cloud.replaceFixedFields(filename)
        time.sleep(1)

        # 3. Replace custom fields
        print("3) Replacing custom fields...")
        cloud.replaceCustomFields(filename)
        time.sleep(1)

        # 4. Replace status IDs
        print("4) Replacing status IDs...")
        cloud.replaceStatus(filename, [
            'destinationStatus": {"type": "ID", "value": "',
            'toStatus": [{"type": "ID", "value": "',
            'fromStatus": [{"type": "ID", "value": "'
        ])
        time.sleep(1)

        # 5. Replace priority IDs
        print("5) Replacing priority IDs...")
        cloud.replacePriority(filename, [
            'fieldType": "priority", "type": "SET", "value": {"type": "ID", "value": "',
            'selectedFieldType": "priority", "comparison": "EQUALS", "compareValue": {"type": "ID", "value": "'
        ])
        time.sleep(1)

        # 6. Replace issue type IDs
        print("6) Replacing issue type IDs...")
        cloud.replaceIssueType(filename, [
            'fieldType": "issuetype", "type": "SET", "value": {"type": "ID", "value": "',
            'selectedFieldType": "issuetype", "comparison": "EQUALS", "compareValue": {"type": "ID", "value": "'
        ])
        time.sleep(1)

        # 7. Replace project IDs
        print("7) Replacing project IDs...")
        cloud.replaceProject(filename, [
            'fieldType": "project", "type": "SET", "value": {"type": "ID", "value": "',
            'selectedFieldType": "project", "comparison": "EQUALS", "compareValue": {"type": "ID", "value": "',
            'projectId": "'
        ])
        time.sleep(1)

        # 8. Replace users
        print("8) Replacing users...")
        cloud.replaceUsers(filename, [
            '"authorAccountId": "',
            '"actorAccountId": "',
            '{"type": "ID", "value": "'
        ])
        time.sleep(1)

        # 9. Replace JIRAUSER references
        print("9) Replacing JIRAUSER users...")
        cloud.replaceJIRAUSERUsers(filename)
        time.sleep(1)

        # 10. Replace IDs in ONE_OF/NOT_ONE_OF conditions
        print("10) Processing ONE_OF/NOT_ONE_OF conditions...")
        cloud.replaceOneOfCondition(filename, 'status')
        cloud.replaceOneOfCondition(filename, 'issuetype')
        cloud.replaceOneOfCondition(filename, 'priority')
        time.sleep(1)

        # 11. Format final JSON
        print("11) Formatting final file...")
        cloud.formatJSON(filename)

        # 12. Split files if requested
        if splitFiles:
            print("\n12) Creating separate files for each rule...")
            with open(filename + "-modified-for-cloud.json", "r", encoding='utf8') as f:
                data = json.load(f)

            if "rules" not in data:
                print("\n'rules' field is missing in the JSON. Could not create separate files.")
            else:
                i = 0
                for i, automation in enumerate(data['rules'], start=1):
                    if automation.get('state') != 'DISABLED':
                        automation['name'] = f"{i} - {automation.get('name')}"
                        new_filename = f"{i}-{automation.get('id')}-modified-for-cloud.json"

                        with open(new_filename, 'w', encoding='utf8') as out_json_file:
                            json.dump({"rules": [automation], "cloud": False}, out_json_file, indent=2)
                        print(f"   File created: {new_filename}")

        # 13. Generate a final mapping Excel with server->cloud IDs
        print("\n13) Generating mapping Excel file...")
        cloud.generateMappingExcel(mapping_data, "mapping_result.xlsx")

        print("\nMigration completed successfully!")
        print(f"Main file: {filename}-modified-for-cloud-pretty.json")
        print("Mapping file: mapping_result.xlsx")

    except Exception as e:
        print(f"\nError during migration: {str(e)}")

if __name__ == "__main__":
    main()
