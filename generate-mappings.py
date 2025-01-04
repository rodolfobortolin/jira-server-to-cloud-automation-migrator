#!/usr/bin/env python3

import sys
import time
import logging
import requests
from tqdm import tqdm
from openpyxl import Workbook
from requests.auth import HTTPBasicAuth

# -------------------------------------------------------------------
# Logging configuration
# -------------------------------------------------------------------
logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(message)s')

# Console logs: show progress bars and high-level info
stdout_handler = logging.StreamHandler(sys.stdout)
stdout_handler.setLevel(logging.INFO)
stdout_handler.setFormatter(formatter)
logger.addHandler(stdout_handler)

# -------------------------------------------------------------------
# Jira Data Center credentials and base URL
# -------------------------------------------------------------------
JIRA_URL = "http://localhost:8080"
USERNAME = "rodolfobortolin"
PASSWORD = "admin"

# -------------------------------------------------------------------
# Jira request helper
# -------------------------------------------------------------------
def get_data_from_jira(endpoint):
    """
    Sends a GET request to Jira (Data Center) for the specified endpoint.
    Logs the endpoint being requested.
    """
    full_url = f"{JIRA_URL}/rest/api/2/{endpoint}"
    logger.info(f"> Fetching data from: {full_url}")

    response = requests.get(
        full_url,
        auth=HTTPBasicAuth(USERNAME, PASSWORD),
        verify=False  # WARNING: only use verify=False in non-production environments
    )
    response.raise_for_status()
    return response.json()

# -------------------------------------------------------------------
# Excel creation function
# -------------------------------------------------------------------
def create_mapping_excel():
    logger.info("\nStarting the creation of 'mapping.xlsx'...")

    wb = Workbook()

    # 1) Users
    logger.info("\n[1/7] Fetching and writing user data...")
    users_sheet = wb.active
    users_sheet.title = "users"
    users_sheet.append(["user_name", "lower_email_address"])

    # We use a generic '.' search to match all, up to 1000 results
    users = get_data_from_jira("user/search?username=.&maxResults=1000")
    logger.info(f"   Found {len(users)} users. Writing them now...")
    time.sleep(1)  # Just to simulate a short pause
    for user in tqdm(users, desc="Writing user data", ncols=100):
        users_sheet.append([
            user.get('key', ''),
            (user.get('emailAddress', '') or '').lower()
        ])

    # 2) Custom Fields
    logger.info("\n[2/7] Fetching and writing custom fields...")
    cf_sheet = wb.create_sheet("customFields")
    cf_sheet.append(["id", "cfname"])

    fields = get_data_from_jira("field")
    custom_fields = [f for f in fields if f.get('custom', False)]
    logger.info(f"   Found {len(custom_fields)} custom fields. Writing them now...")
    time.sleep(1)
    for field in tqdm(custom_fields, desc="Writing custom fields", ncols=100):
        field_id = field['id'].replace('customfield_', '')
        cf_sheet.append([field_id, field['name']])

    # 3) Projects
    logger.info("\n[3/7] Fetching and writing projects...")
    proj_sheet = wb.create_sheet("projects")
    proj_sheet.append(["id", "pkey"])

    projects = get_data_from_jira("project")
    logger.info(f"   Found {len(projects)} projects. Writing them now...")
    time.sleep(1)
    for project in tqdm(projects, desc="Writing projects", ncols=100):
        proj_sheet.append([project['id'], project['key']])

    # 4) Status
    logger.info("\n[4/7] Fetching and writing statuses...")
    status_sheet = wb.create_sheet("status")
    status_sheet.append(["id", "pname"])

    statuses = get_data_from_jira("status")
    logger.info(f"   Found {len(statuses)} statuses. Writing them now...")
    time.sleep(1)
    for status in tqdm(statuses, desc="Writing status", ncols=100):
        status_sheet.append([status['id'], status['name']])

    # 5) Priority
    logger.info("\n[5/7] Fetching and writing priorities...")
    prio_sheet = wb.create_sheet("priority")
    prio_sheet.append(["id", "pname"])

    priorities = get_data_from_jira("priority")
    logger.info(f"   Found {len(priorities)} priorities. Writing them now...")
    time.sleep(1)
    for priority in tqdm(priorities, desc="Writing priorities", ncols=100):
        prio_sheet.append([priority['id'], priority['name']])

    # 6) Issue Type
    logger.info("\n[6/7] Fetching and writing issue types...")
    type_sheet = wb.create_sheet("issuetype")
    type_sheet.append(["id", "pname"])

    types = get_data_from_jira("issuetype")
    logger.info(f"   Found {len(types)} issue types. Writing them now...")
    time.sleep(1)
    for issuetype in tqdm(types, desc="Writing issue types", ncols=100):
        type_sheet.append([issuetype['id'], issuetype['name']])

    # 7) Resolutions
    logger.info("\n[7/7] Fetching and writing resolutions...")
    res_sheet = wb.create_sheet("resolutions")
    res_sheet.append(["id", "pname"])

    resolutions = get_data_from_jira("resolution")
    logger.info(f"   Found {len(resolutions)} resolutions. Writing them now...")
    time.sleep(1)
    for resolution in tqdm(resolutions, desc="Writing resolutions", ncols=100):
        res_sheet.append([resolution['id'], resolution['name']])

    # Save the workbook
    output_file = "mapping.xlsx"
    wb.save(output_file)
    logger.info(f"\n✔ Successfully created '{output_file}' with all mappings.")

# -------------------------------------------------------------------
# Main entry point
# -------------------------------------------------------------------
if __name__ == "__main__":
    try:
        create_mapping_excel()
    except Exception as e:
        logger.error(f"Error: {e}")
